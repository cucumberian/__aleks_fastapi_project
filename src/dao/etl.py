import pandas as pd
from sqlalchemy import text
from datetime import date
from db.database import async_session_maker
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import asyncio


async def process_excel_and_insert_data(file_path):
    """
    Обрабатывает данные из Excel файла и вставляет их в таблицы old_apart и new_apart.
    """

    async def insert_data(table_name, data):
        """
        Вставляет данные в указанную таблицу с обработкой конфликтов.
        """
        async with async_session_maker() as session:
            try:
                for row_data in data:
                    query = text(f"""
                        INSERT INTO recommendation.{table_name} ({','.join(row_data.keys())}, insert_date)
                        VALUES ({','.join([':' + key for key in row_data.keys()])}, NOW())
                        ON CONFLICT {'(fio, house_address, apart_number)' if table_name == 'old_apart' else '(unique_id)'} DO UPDATE SET
                            {', '.join([key + ' = EXCLUDED.' + key for key in row_data.keys()])},
                            insert_date = NOW()
                    """)
                    await session.execute(query, row_data)
                await session.commit()
            except Exception as e:
                await session.rollback()
                print(f"Ошибка при вставке данных в {table_name}: {e}")
                return False
            return True

    # Обработка листа "old_apart"
    df_old_apart = pd.read_excel(file_path, sheet_name='old_apart')

    # Преобразование типов данных для old_apart
    df_old_apart['№ кв-ры'] = pd.to_numeric(df_old_apart['№ кв-ры'], errors='coerce').fillna(0).astype(int)
    df_old_apart['кол-во комнат'] = pd.to_numeric(df_old_apart['кол-во комнат'], errors='coerce').fillna(0).astype(int)
    df_old_apart['площ. жил. пом.'] = pd.to_numeric(df_old_apart['площ. жил. пом.'], errors='coerce').fillna(0.0).astype(float)
    df_old_apart['общ. пл.'] = pd.to_numeric(df_old_apart['общ. пл.'], errors='coerce').fillna(0.0).astype(float)
    df_old_apart['жил. пл.'] = pd.to_numeric(df_old_apart['жил. пл.'], errors='coerce').fillna(0.0).astype(float)
    df_old_apart['Кол-во членов семьи'] = pd.to_numeric(df_old_apart['Кол-во членов семьи'], errors='coerce').fillna(0).astype(int)
    df_old_apart['Потребность'] = pd.to_numeric(df_old_apart['Потребность'], errors='coerce').fillna(0).astype(int)
    df_old_apart['мин этаж'] = pd.to_numeric(df_old_apart['мин этаж'], errors='coerce').fillna(0).astype(int)
    df_old_apart['макс этаж'] = pd.to_numeric(df_old_apart['макс этаж'], errors='coerce').fillna(0).astype(int)
    df_old_apart['адрес дома'] = df_old_apart['адрес дома'].astype(str)


    old_apart_data = [{
        'district': row['Округ'],
        'area': row['район'],
        'apart_number': str(int(row['№ кв-ры'])),
        'type_of_settlement': row['Вид засел.'],
        'room_count': int(row['кол-во комнат']),
        'full_living_area': float(row['площ. жил. пом.']),
        'total_living_area': float(row['общ. пл.']),
        'living_area': float(row['жил. пл.']),
        'house_address': row['адрес дома'],
        'fio': row['ФИО'],
        'members_amount': row['Кол-во членов семьи'],
        'need': row['Потребность'],
        'min_floor': int(row['мин этаж']),
        'max_floor': int(row['макс этаж'])
    } for _, row in df_old_apart.iterrows()]

    if not await insert_data('old_apart', old_apart_data):
        return False

    # Обработка листа "new_apart"
    df_new_apart = pd.read_excel(file_path, sheet_name='new_apart')
    df_new_apart['К_Инв/к'] = df_new_apart['К_Инв/к'].apply(lambda x: 1 if x == 'да' else 0).astype(int)
    df_new_apart = df_new_apart[df_new_apart['К_Состояние'] == 'Свободная']

    # Преобразование типов данных для new_apart
    df_new_apart['Адрес_№ кв'] = pd.to_numeric(df_new_apart['Адрес_№ кв'], errors='coerce').fillna(0).astype(int)
    df_new_apart['К_Этаж'] = pd.to_numeric(df_new_apart['К_Этаж'], errors='coerce').fillna(0).astype(int)
    df_new_apart['К_Комн'] = pd.to_numeric(df_new_apart['К_Комн'], errors='coerce').fillna(0).astype(int)
    df_new_apart['Площадь общая'] = pd.to_numeric(df_new_apart['Площадь общая'], errors='coerce').fillna(0.0).astype(float)
    df_new_apart['Площадь общая(б/л)'] = pd.to_numeric(df_new_apart['Площадь общая(б/л)'], errors='coerce').fillna(0.0).astype(float)
    df_new_apart['Площадь жилая'] = pd.to_numeric(df_new_apart['Площадь жилая'], errors='coerce').fillna(0.0).astype(float)
    df_new_apart['Дата переселения'] = date.today()
    df_new_apart['Сл.инф_APART_ID'] = pd.to_numeric(df_new_apart['Сл.инф_APART_ID'], errors='coerce').fillna(0).astype(int)

    new_apart_data = [{
        'district': row['Адрес_Округ'],
        'area': row['Адрес_Мун.округ'],
        'house_address': row['Адрес_Короткий'],
        'apart_number': str(row['Адрес_№ кв']),
        'floor': row['К_Этаж'],
        'room_count': row['К_Комн'],
        'full_living_area': row['Площадь общая'],
        'total_living_area': row['Площадь общая(б/л)'],
        'living_area': row['Площадь жилая'],
        'status_marker': row['К_Инв/к'],
        'unique_id': row['Сл.инф_APART_ID']
    } for _, row in df_new_apart.iterrows()]

    if not await insert_data('new_apart', new_apart_data):
        return False

    return True

async def save_views_to_excel(output_filename, output_path, selected_districts=None, old_selected_area=None,
                              new_selected_area=None, date=False, new_selected_addresses=None,
                              old_selected_addresses=None):
    """Сохраняет данные из представлений в Excel файл."""
    try:
        views = [
            'new_apart_all_last', 'res_of_rec_last',
            'ranked_last', 'where_not_last'
        ] if date else [
            'new_apart_all', 'result_of_recommendation',
            'ranked_with_district', 'where_not_offered'
        ]

        async with async_session_maker() as session:
            wb = Workbook()

            for view in views:
                print(f"Обработка представления: {view}")

                if view == 'ranked_with_district':
                    # --- Запрос для ранжирования квартир ---
                    ranked_query = """
                        WITH RankedNewApartments AS (
                            SELECT 
                                new_apart_id, room_count, living_area, full_living_area, total_living_area, district, area, house_adress,
                                ROW_NUMBER() OVER (PARTITION BY room_count, district ORDER BY living_area ASC, full_living_area ASC, total_living_area ASC) as rank_num
                            FROM recommendation.new_apart
                            WHERE 1=1 
                    """

                    ranked_params = {}

                    if selected_districts:
                        ranked_query += " AND district = ANY(:selected_districts)"
                        ranked_params['selected_districts'] = selected_districts

                    if new_selected_area:
                        ranked_query += " AND area = ANY(:new_selected_area)"
                        ranked_params['new_selected_area'] = new_selected_area

                    if new_selected_addresses:
                        ranked_query += " AND house_adress = ANY(:new_selected_addresses)"
                        ranked_params['new_selected_addresses'] = new_selected_addresses

                    if date:
                        ranked_query += " AND insert_date = (SELECT MAX(insert_date) FROM recommendation.new_apart)"

                    ranked_query += """
                        ),
                        RankedOldApartments AS (
                            SELECT 
                                old_apart_id, room_count, living_area, full_living_area, total_living_area, need, district, area, house_adress,
                                ROW_NUMBER() OVER (PARTITION BY room_count, district ORDER BY living_area ASC, full_living_area ASC, total_living_area ASC) as rank_num
                            FROM recommendation.old_apart
                            WHERE old_apart_id NOT IN (SELECT old_apart_id FROM recommendation.offer) 
                    """

                    if selected_districts:
                        ranked_query += " AND district = ANY(:selected_districts)"

                    if old_selected_area:
                        ranked_query += " AND area = ANY(:old_selected_area)"
                        ranked_params['old_selected_area'] = old_selected_area

                    if old_selected_addresses:
                        ranked_query += " AND house_adress = ANY(:old_selected_addresses)"
                        ranked_params['old_selected_addresses'] = old_selected_addresses

                    if date:
                        ranked_query += " AND insert_date = (SELECT MAX(insert_date) FROM recommendation.old_apart)"

                    ranked_query += """
                        )
                        SELECT 
                            rna.new_apart_id, rna.room_count, rna.living_area, rna.full_living_area, rna.total_living_area, rna.district, rna.area, rna.house_adress, rna.rank_num,
                            roa.old_apart_id, roa.need
                        FROM RankedNewApartments rna
                        LEFT JOIN RankedOldApartments roa ON rna.room_count = roa.room_count AND rna.district = roa.district AND rna.rank_num = roa.rank_num
                    """

                    result = await session.execute(text(ranked_query), ranked_params)
                    df = pd.DataFrame(result.fetchall(), columns=result.keys())

                    # Сохранение в Excel
                    if 'Ранг' not in wb.sheetnames:
                        ws = wb.create_sheet('Ранг')
                    else:
                        ws = wb['Ранг']

                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)

                elif view == 'ranked_last':
                    # --- Запрос для ранжирования квартир (только последние данные) ---
                    ranked_last_query = """
                        WITH RankedNewApartments AS (
                            SELECT 
                                new_apart_id, room_count, living_area, full_living_area, total_living_area, district, area, house_address,
                                ROW_NUMBER() OVER (PARTITION BY room_count, district ORDER BY living_area ASC, full_living_area ASC, total_living_area ASC) as rank_num
                            FROM recommendation.new_apart
                            WHERE insert_date = (SELECT MAX(insert_date) FROM recommendation.new_apart) 
                    """

                    ranked_last_params = {}

                    if selected_districts:
                        ranked_last_query += " AND district = ANY(:selected_districts)"
                        ranked_last_params['selected_districts'] = selected_districts

                    if new_selected_area:
                        ranked_last_query += " AND area = ANY(:new_selected_area)"
                        ranked_last_params['new_selected_area'] = new_selected_area

                    if new_selected_addresses:
                        ranked_last_query += " AND house_adress = ANY(:new_selected_addresses)"
                        ranked_last_params['new_selected_addresses'] = new_selected_addresses

                    ranked_last_query += """
                        ),
                        RankedOldApartments AS (
                            SELECT 
                                old_apart_id, room_count, living_area, full_living_area, total_living_area, need, district, area, house_adress,
                                ROW_NUMBER() OVER (PARTITION BY room_count, district ORDER BY living_area ASC, full_living_area ASC, total_living_area ASC) as rank_num
                            FROM recommendation.old_apart
                            WHERE old_apart_id NOT IN (SELECT old_apart_id FROM recommendation.offer) 
                            AND insert_date = (SELECT MAX(insert_date) FROM recommendation.old_apart)
                    """

                    if selected_districts:
                        ranked_last_query += " AND district = ANY(:selected_districts)"

                    if old_selected_area:
                        ranked_last_query += " AND area = ANY(:old_selected_area)"
                        ranked_last_params['old_selected_area'] = old_selected_area

                    if old_selected_addresses:
                        ranked_last_query += " AND house_adress = ANY(:old_selected_addresses)"
                        ranked_last_params['old_selected_addresses'] = old_selected_addresses

                    ranked_last_query += """
                        )
                        SELECT 
                            rna.new_apart_id, rna.room_count, rna.living_area, rna.full_living_area, rna.total_living_area, rna.district, rna.area, rna.house_adress, rna.rank_num,
                            roa.old_apart_id, roa.need
                        FROM RankedNewApartments rna
                        LEFT JOIN RankedOldApartments roa ON rna.room_count = roa.room_count AND rna.district = roa.district AND rna.rank_num = roa.rank_num
                    """

                    result = await session.execute(text(ranked_last_query), ranked_last_params)
                    df = pd.DataFrame(result.fetchall(), columns=result.keys())

                    # Сохранение в Excel
                    if 'Ранг (последние данные)' not in wb.sheetnames:
                        ws = wb.create_sheet('Ранг (последние данные)')
                    else:
                        ws = wb['Ранг (последние данные)']

                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)

                else:
                    query = text(f"SELECT * FROM recommendation.{view}")
                    params = {}

                    if view == 'new_apart_all':
                        if selected_districts:
                            query += " WHERE district = ANY(:selected_districts)"
                            params['selected_districts'] = selected_districts

                        if new_selected_area:
                            query += " AND area = ANY(:new_selected_area)"
                            params['new_selected_area'] = new_selected_area

                        if new_selected_addresses:
                            query += " AND house_adress = ANY(:new_selected_addresses)"
                            params['new_selected_addresses'] = new_selected_addresses

                        if date:
                            query += " AND insert_date = (SELECT MAX(insert_date) FROM recommendation.new_apart)"

                    elif view == 'where_not_offered':
                        query += " WHERE 1=1"
                        if selected_districts:
                            query += " AND district = ANY(:selected_districts)"
                            params['selected_districts'] = selected_districts

                        if old_selected_area:
                            query += " AND area = ANY(:old_selected_area)"
                            params['old_selected_area'] = old_selected_area

                        if old_selected_addresses:
                            query += " AND house_address = ANY(:old_selected_addresses)"
                            params['old_selected_addresses'] = old_selected_addresses

                        if date:
                            query += " AND insert_date = (SELECT MAX(insert_date) FROM recommendation.old_apart)"

                    elif view == 'result_of_recommendation':
                        query += " WHERE 1=1"
                        if selected_districts:
                            query += " AND district = ANY(:selected_districts)"
                            params['selected_districts'] = selected_districts

                        if old_selected_area:
                            query += " AND area_old = ANY(:old_selected_area)"
                            params['old_selected_area'] = old_selected_area

                        if new_selected_area:
                            query += " AND area_new = ANY(:new_selected_area)"
                            params['new_selected_area'] = new_selected_area

                        if old_selected_addresses:
                            query += " AND house_address_old = ANY(:old_selected_addresses)"
                            params['old_selected_addresses'] = old_selected_addresses

                        if new_selected_addresses:
                            query += " AND house_address_new = ANY(:new_selected_addresses)"
                            params['new_selected_addresses'] = new_selected_addresses

                        if date:
                            query += " AND insert_date = (SELECT MAX(insert_date) FROM recommendation.result_of_recommendation)"

                    print(f"Выполнение запроса для представления {view}: {query}")

                    try:
                        result = await session.execute(query, params)
                        df = pd.DataFrame(result.fetchall(), columns=result.keys())
                        df = df.dropna(how='all')

                        if view not in wb.sheetnames:
                            ws = wb.create_sheet(view)
                        else:
                            ws = wb[view]

                        for r in dataframe_to_rows(df, index=False, header=True):
                            ws.append(r)

                    except Exception as e:
                        print(f"Ошибка выполнения запроса для представления {view}: {e}")

            # Удаление дефолтного листа
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            wb.save(f"{output_path}{output_filename}.xlsx")
            return f"{output_path}{output_filename}.xlsx"

    except Exception as e:
        print(f"Ошибка: {e}")


async def generate_data_tables(old_selected_districts=None, new_selected_areas=None,
                        old_selected_areas=None, date=None,
                        new_selected_addresses=None, old_selected_addresses=None):
    """Генерирует таблицы данных."""
    tables = []

    try:
        async with async_session_maker() as session:
            for view in ['Ранг']:
                query_old = """
                    SELECT old_apart_id, room_count, living_area, full_living_area, total_living_area, district, area, house_address, need
                    FROM recommendation.old_apart WHERE 1=1
                """
                query_new = """
                    SELECT new_apart_id, room_count, living_area, full_living_area, total_living_area, district, area, house_address, status_marker
                    FROM recommendation.new_apart WHERE 1=1
                """

                old_apart_query_params = {}
                new_apart_query_params = {}

                if old_selected_districts:
                    query_old += " AND district = ANY(:selected_districts)"
                    old_apart_query_params['selected_districts'] = old_selected_districts
                    query_new += " AND district = ANY(:selected_districts)"
                    new_apart_query_params['selected_districts'] = old_selected_districts

                if new_selected_areas:
                    query_new += " AND area = ANY(:new_selected_areas)"
                    new_apart_query_params['new_selected_areas'] = new_selected_areas

                if old_selected_areas:
                    query_old += " AND area = ANY(:old_selected_areas)"
                    old_apart_query_params['old_selected_areas'] = old_selected_areas

                if old_selected_addresses:
                    query_old += " AND house_address = ANY(:old_selected_addresses)"
                    old_apart_query_params['old_selected_addresses'] = old_selected_addresses

                if new_selected_addresses:
                    query_new += " AND house_address = ANY(:new_selected_addresses)"
                    new_apart_query_params['new_selected_addresses'] = new_selected_addresses

                if date:
                    query_new += " AND insert_date = (SELECT MAX(insert_date) FROM recommendation.new_apart)"
                    query_old += " AND insert_date = (SELECT MAX(insert_date) FROM recommendation.old_apart)"

                result_old = await session.execute(text(query_old), old_apart_query_params)
                df_old = pd.DataFrame(result_old.fetchall(), columns=result_old.keys())

                result_new = await session.execute(text(query_new), new_apart_query_params)
                df_new = pd.DataFrame(result_new.fetchall(), columns=result_new.keys())

                # Создаем комбинированный столбец для новых и старых квартир
                df_new['combined_area'] = list(
                    zip(df_new['living_area'], df_new['full_living_area'], df_new['total_living_area']))
                df_old['combined_area'] = list(
                    zip(df_old['living_area'], df_old['full_living_area'], df_old['total_living_area']))

                # Сортируем DataFrame по каждому из значений в combined_area для новых квартир
                df_new = df_new.sort_values(['living_area', 'full_living_area', 'total_living_area'],
                                            ascending=True)

                # Присваиваем ранги новым квартирам
                df_new['rank'] = df_new.groupby(['room_count', 'district'])['combined_area'].rank(
                    method='dense').astype(int)

                # Инициализируем колонку для рангов в df_old
                df_old['rank'] = 0  # Инициализация колонки для рангов

                # Создаем словарь для хранения максимальных рангов по количеству комнат
                max_rank_by_room_count = df_new.groupby('room_count')['rank'].max().to_dict()

                # Инициализация списка для старых квартир, которым не найдется закрывающая новая квартира
                unmatched_old = []

                # Присваиваем ранги старым квартирам на основе новых
                for idx, old_row in df_old.iterrows():
                    filtered_new = df_new[
                        (df_new['room_count'] == old_row['room_count']) &
                        (df_new['district'] == old_row['district']) &
                        (df_new['living_area'] >= old_row['living_area']) &
                        (df_new['full_living_area'] >= old_row['full_living_area']) &
                        (df_new['total_living_area'] >= old_row['total_living_area']) &
                        (df_new['status_marker'] == old_row['need'])
                    ]

                    if not filtered_new.empty:
                        min_rank = filtered_new['rank'].min()
                        df_old.at[idx, 'rank'] = min_rank
                    else:
                        unmatched_old.append(idx)

                # Присваиваем максимальный ранг + 1 для старых квартир, которые не были закрыты новыми
                for room_count in df_old['room_count'].unique():
                    max_rank_new = max_rank_by_room_count.get(room_count, 0)
                    df_old.loc[(df_old['rank'] == 0) & (df_old['room_count'] == room_count), 'rank'] = max_rank_new + 1

                # Объединяем данные старых и новых квартир
                df_combined = pd.concat([df_old.assign(status='old'), df_new.assign(status='new')], ignore_index=True)

                # Присваивание групп рангов
                df_combined['rank_group'] = df_combined['rank'].astype(int)

                # Группировка и расчет показателей
                df = df_combined.groupby(['room_count', 'rank_group']).agg(
                    Пот_ть=('old_apart_id', 'count'),
                    Ресурс=('new_apart_id', 'count')
                ).reset_index()

                df['Баланс'] = df['Ресурс'] - df['Пот_ть']

                # Добавление итоговых данных
                def add_totals(df, max_rank_by_room_count):
                    total_potency = 0
                    total_resource = 0
                    total_balance = 0

                    new_rows = []
                    previous_row = None
                    start_rank = None

                    for i in range(len(df)):
                        row = df.iloc[i].to_dict()
                        current_rank = row['rank_group']
                        room_count = row['room_count']
                        max_rank = max_rank_by_room_count.get(room_count, 0) + 1

                        if previous_row is not None:
                            if isinstance(previous_row['rank_group'], str) and '-' in previous_row['rank_group']:
                                previous_rank = int(previous_row['rank_group'].split('-')[-1])
                            else:
                                previous_rank = previous_row['rank_group']

                            if ((previous_row['Пот_ть'] == 0) or (row['Пот_ть'] == 0)) and current_rank != max_rank and previous_rank != max_rank:
                                previous_row['Пот_ть'] += row['Пот_ть']
                                previous_row['Ресурс'] += row['Ресурс']
                                previous_row['Баланс'] += row['Баланс']
                                previous_row['rank_group'] = f"{start_rank}-{current_rank}"
                            else:
                                if previous_row is not None:
                                    new_rows.append(previous_row)

                                previous_row = row
                                start_rank = current_rank
                        else:
                            previous_row = row
                            start_rank = current_rank

                    if previous_row is not None:
                        new_rows.append(previous_row)

                    df_new = pd.DataFrame(new_rows)
                    total_potency = df_new['Пот_ть'].sum()
                    total_resource = df_new['Ресурс'].sum()
                    total_balance = df_new['Баланс'].sum()

                    totals = pd.DataFrame([{
                        'rank_group': 'Итог',
                        'Пот_ть': total_potency,
                        'Ресурс': total_resource,
                        'Баланс': total_balance
                    }])

                    df_with_totals = pd.concat([df_new, totals], ignore_index=True)

                    return df_with_totals

                result_data = []
                for room in df['room_count'].unique():
                    room_df = df[df['room_count'] == room].copy()
                    grouped_df = add_totals(room_df, max_rank_by_room_count)
                    grouped_df['room_count'] = room
                    result_data.append(grouped_df)

                df_grouped = pd.concat(result_data, ignore_index=True)

                if df_grouped.empty:
                    print(f"Представление {view} не вернуло данных.")
                    continue

                for room in df_grouped['room_count'].unique():
                    room_df = df_grouped[df_grouped['room_count'] == room].copy()
                    room_df = room_df[['rank_group', 'Пот_ть', 'Ресурс', 'Баланс']]
                    room_df.columns = ['Тип кв.', 'Пот-ть', 'Ресурс', 'Баланс']

                    room_df[['Пот-ть', 'Ресурс', 'Баланс']] = room_df[['Пот-ть', 'Ресурс', 'Баланс']].astype(int)
                    tables.append({
                        'title': f'{room} комната(ы)',
                        'table_html': room_df.to_html(classes='table table-bordered table-hover', index=False)
                    })

    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")

    return tables

"""async def main():
    excel_file_path = await save_views_to_excel(
        output_filename='done', 
        output_path='src/', 
        new_selected_addresses=["ул. Давыдковская, д. 10"],
        old_selected_addresses=["ул. Давыдковская, д. 10, корп. 4", "ул. Кременчугская, д. 5, корп. 1", "ул. Давыдковская, д. 12, корп. 1", "Славянский бул., д. 9, корп. 3"]
    )
    await generate_data_tables(excel_file_path)  # Передаем путь к файлу Excel

if __name__ == '__main__':
    asyncio.run(main())

"""